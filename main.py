import cv2
import threading
from datetime import date
import sqlite3
import flet as ft
import os
import time
import pyttsx3 as pt
import os
def say():
    pt.speak('system initialized')
    pt.speak('welcome to the smart system of a smart coaching  ')
    pt.speak('I am opening the smart attendence application very soon students ')
threading.Thread(target=say,daemon=True).start()
import openpyxl
from deepface import DeepFace

import flet_charts as fch
class Attendence_system():
    def __init__(self,excel='attendence.xlsx',cam=0):
        self.workbook=openpyxl.Workbook()
        self.excell=excel
        self.worksheet=self.workbook.active
        self.cam=cam
        self.face_detector=cv2.CascadeClassifier(cv2.data.haarcascades+"haarcascade_frontalface_default.xml")
        self.conn=sqlite3.connect('attendence.db',check_same_thread=False)
        self.cursor=self.conn.cursor()
        self.cursor.execute('''CREATE TABLE IF NOT EXISTS students(id INTEGER PRIMARY KEY AUTOINCREMENT,path TEXT,name Text)''')
        self.cursor.execute('''CREATE TABLE IF NOT EXISTS attendence(id INTEGER,name TEXT,mark Text,date Text)''')
        self.conn.commit()
    
    def attend_worker(self,today,face,totalid,func):
                            id=1
                            threading.Thread(target=func,args=(face,),daemon=True).start()
                            conn=sqlite3.connect('attendence.db')
                            cursor=conn.cursor()
                            while id<=totalid:
                              cursor.execute(f'''SELECT PATH FROM STUDENTS WHERE ID="{id}"''')
                              path=cursor.fetchone()
                              if path:
                                 path=path[0]
                                 try:
                                   result=DeepFace.verify(face,path)
                                   result=result['verified']
                                   if result:
                                       cursor.execute(f'''select name from attendence where id="{id}" and date="{today}"''')
                                       namee=cursor.fetchone()
                                       if namee :break
                                       else:
                                           cursor.execute(f'''SELECT name FROM STUDENTS WHERE ID="{id}"''')
                                           name=cursor.fetchone()
                                           if name:
                                            name=name[0]
                                            cursor.execute('''INSERT INTO attendence (id,name,mark,date) values (?,?,?,?)''',(id,name,'present',today))
                                            conn.commit()
                                            print('attendence made')
        
                                 except Exception as e:
                                        print(e)
                                        id+=1                
    def attendence(self):
        def show(image):
            while True:
              cv2.imshow('attendence',image)
              if cv2.waitKey(2000)==ord('f'):
                 cv2.destroyAllWindows()
                 break
              break

        self.cursor.execute('''SELECT COUNT(id) from students''')
        total_id=self.cursor.fetchone()
        if total_id:
            total_id=total_id[0]
        else:total_id=0
        today=date.today()
        cam=cv2.VideoCapture(self.cam)
        s,image=cam.read()
        if s:
          faces=self.face_detector.detectMultiScale(image,1.4)
          if len(faces)>0:
              

              for face in faces:
                      x,y,w,h=face
                      face=image[y-30:y+h+30,x-30:x+w+30]
                      threading.Thread(target=self.attend_worker,args=(str(today),face,total_id,show),daemon=True).start()
                      
              id=1
        cam.release()
    def registor(self,name):
       path=name+'.png'
       while True:
        cam=cv2.VideoCapture(0)
        s,image=cam.read()
        if s: 
            faces=self.face_detector.detectMultiScale(image,1.4)
            for face in faces:
                x,y,w,h=face
                im=cv2.rectangle(image,(x,y),(x+w,y+h),(255,0,0))  
                face=image[y-30:y+h+30,x-30:x+w+30]
                cv2.imshow('register',im)
                if cv2.waitKey(1)==ord('a'):
                    cv2.imwrite(path,face)
                    self.cursor.execute(f'''select id from students where name="{name}" and path="{path}"''')
                    if (self.cursor.fetchone()) ==None:
                         self.cursor.execute('''INSERT INTO STUDENTS(path,name) values (?,?)''',(path,name))
                         self.conn.commit()
                         cv2.destroyAllWindows()
                         cam.release()
                         return None
                    else:
                        print('try again with a new username')
                        cv2.destroyAllWindows()
                        cam.release()
                        return 0
    def excel(self,datee=str(date.today())):
        if datee:
         self.cursor.execute(f'''select * from attendence where date="{datee}"''')
         data=self.cursor.fetchall()
        else:
            self.cursor.execute(f'''select * from attendence''')
            data=self.cursor.fetchall()
            datee='all_ever'
        if data:
            self.worksheet.delete_cols(1,self.worksheet.max_column)
            self.worksheet.delete_rows(1,self.worksheet.max_row)
            self.workbook.save(self.excell)
            self.worksheet.append(['ROLL','NAME','ATTENDENCE','DATE'])
            for i in data:
                self.worksheet.append(i)
            self.workbook.save(datee+'.xlsx')
         
        else :
            self.worksheet.delete_cols(1,self.worksheet.max_column)
            self.worksheet.delete_rows(1,self.worksheet.max_row)
            self.worksheet.append(['nothing yet master'])
            self.workbook.save(self.excell)
    def total(self):
        today=str(date.today())
        self.cursor.execute('''SELECT COUNT(*) from students''')
        total=self.cursor.fetchone()
        if total:
            total=total[0]
        else: total=0
        self.cursor.execute(f'''SELECT COUNT(*) from attendence where date="{today}"''')
        total_present=self.cursor.fetchone()
        if total_present:
         total_present=total_present[0]
        else: total_present=0
        return total,total_present
a=Attendence_system(cam=0)
today=(str(date.today()))
with open ('status.txt','r')as f:
    end_date=f.read()

def wizard(page,upper,registor):
    upper.disabled=True
    registor.disabled=True
    global a
    def run(o,a,upper):
        if o.value=='TODAY':
            a.excel()
        else:
            a.excel(datee=None)
        upper.disabled=False
        registor.disabled=False
        page.remove(options,button)


    options=ft.Dropdown(hint_text='DATE',options=[ft.DropdownOption('TODAY'),ft.DropdownOption('ALL EVER ATTENDENCE')],border=1,border_color=ft.Colors.BLUE,border_radius=5)
    button=ft.IconButton(icon=ft.Icons.SAVE,on_click=lambda:run(options,a,upper),hover_color=ft.Colors.BLUE_200)

    page.add(options,button)
def attendence_do(switch):
    global a
    while switch.value:
        a.attendence()
def main(page):
    page.clean()
    global attendence_do
    def attendence_start(switch):
        if switch.value:
            threading.Thread(target=attendence_do,daemon=True,args=(switch,)).start()

    def check_and_update(present,absent):
     while True:
      global a
      conn=sqlite3.connect('attendence.db',check_same_thread=False)
      cursor=conn.cursor()
      today=str(date.today())
      cursor.execute('''SELECT COUNT(*) from students''')
      total=cursor.fetchone()
      if total:
            total=total[0]
      else: total=0
      cursor.execute(f'''SELECT COUNT(*) from attendence where date="{today}"''')
      total_present=cursor.fetchone()
      if total_present:
               total_present=total_present[0]
      else: total_present=0
      if total and total_present:
       present.value=(total_present/total)*100
       absent.value=(100-present.value)
       present.title=str(present.value)+'%'
       absent.title=str(absent.value)+'%'
       time.sleep(1)
          
        
    def registeration(registor,upper):
        registor.disabled=True
        upper.disabled=True
        def regis(name,button,registor,upper):
            if name.value:
                x=a.registor(name.value)
                if x==0: name.label='username already exists';name.color='red'
                else: page.remove(name,button);registor.disabled=False;upper.disabled=False
        name=ft.TextField(hint_text='username must be uniqe',label='USERNAME',border=1,border_color=ft.Colors.BLUE,border_radius=5)
        button=ft.IconButton(ft.Icons.SAVE_ALT_OUTLINED,on_click=lambda:regis(name,button,registor,upper),hover_color=ft.Colors.BLUE_200)
        page.add(name,button)
    page.title='ATTENDIFY'
    page.theme_mode=ft.ThemeMode.DARK
    
    upper=ft.AppBar(leading=ft.IconButton(ft.Icons.FILE_COPY,icon_color='blue',focus_color='black',on_click=lambda : wizard(page,upper,registor)),title=ft.Text('ATTENDI.fy',font_family='Algerian'),center_title=True)
    registor=(ft.IconButton(ft.Icons.APP_REGISTRATION,'blue',focus_color='black',on_click=lambda:registeration(registor,upper),hover_color=ft.Colors.BLUE_200))
    present=fch.PieChartSection(value=50,radius=50,color='green')
    absent=fch.PieChartSection(value=50,radius=50,color='red')
    z=fch.PieChart(sections=[present,absent],width=200,height=200,align=ft.Alignment.CENTER)
    attendence=ft.Switch(label='ATTENDENCE',on_change=lambda:attendence_start(attendence),align=ft.Alignment.CENTER,margin=30)
    threading.Thread(target=check_and_update,args=(present,absent),daemon=True).start()

    page.add(upper,registor,attendence,z)

def startup(page):
    global today
    global end_date
    if today>=end_date:
        def pay(key,page):
            if (len(str(key.value)) == 14 and str(key.value)[4]=='D') and ('@ctivate' in str(key.value) or '@king' in str(key.value) ):
                main(page)
                with open('status.txt','w')as f:
                    d=today.split('-')
                    if d[1]=='12':
                        d[1]='01'
                        d[0]=str(int(d[0])+1)
                        d='-'.join(d)
                        f.write(d)
                    else:
                     if len(d[1])==1 and int(d[1])<9:
                      d[1]='0'+str(int(d[1])+1)
                     else:d[1]=str(int(d[1])+1)
                     d='-'.join(d)
                     f.write(d)
            else:
                key.color='red';key.label='contact diivyanshu.space on instagram'
        page.title='PAYMENT DAY'
        page.window.height=600
        page.window.width=500
        page.window.maximizable=False
        page.bgcolor=ft.Colors.BLACK
        qr=ft.Image('payment.png',fade_in_animation=ft.Animation(4500),height=400,width=600)
        page.add(qr)
        activation_key=ft.TextField(label='ACTIVATION KEY',hint_text='subscription expired')
        submit=ft.Button('SUBMIT',on_click=lambda: pay(activation_key,page))
        page.add(activation_key,submit)
    else:
     page.title='ATTENDI.fy'
     page.bgcolor=ft.Colors.BLACK
     def start(page):
             key=ft.TextField(label='ACTIVATION KEY')
             submit=ft.Button('START NOW',on_click=lambda:check(key,page))
             page.add(key,submit)
     def check(key,page):
         if key.value=='start':
             main(page)
     page.clean()
     page.window.height=600
     page.window.width=500
     page.window.maximizable=False
     img=ft.Image('INDUSS.png',fade_in_animation=ft.Animation(4500),height=400,width=600)
     page.add(img)
     start(page)
    


ft.run(startup)



